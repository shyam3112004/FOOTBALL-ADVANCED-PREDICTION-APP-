import io
from datetime import datetime
from typing import Dict, List, Any, Optional

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, Reference


# ─── Color constants ──────────────────────────────────────────────────────────
DARK_GREEN  = "00703C"
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
HEADER_FILL = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
ALT_FILL    = PatternFill(start_color="F2F7F4", end_color="F2F7F4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
HIGH_FILL   = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
MED_FILL    = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
LOW_FILL    = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")


def _header_row(ws, row: int, values: List[str], freeze: bool = False):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _data_row(ws, row: int, values: List[Any], alt: bool = False):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_FILL


def _pct_row(ws, row: int, values: List[Any], alt: bool = False):
    """Same as _data_row but percentage columns formatted."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_FILL
        if isinstance(val, float) and 0 <= val <= 1:
            cell.number_format = "0.00%"


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 32)


def _add_conditional_formatting(ws, col_letter: str, start_row: int, end_row: int):
    """Green >60%, amber 40-60%, red <40% for probability columns."""
    pct_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.conditional_formatting.add(pct_range, CellIsRule(
        operator="greaterThan", formula=["0.6"], fill=HIGH_FILL
    ))
    ws.conditional_formatting.add(pct_range, CellIsRule(
        operator="between", formula=["0.4", "0.6"], fill=MED_FILL
    ))
    ws.conditional_formatting.add(pct_range, CellIsRule(
        operator="lessThan", formula=["0.4"], fill=LOW_FILL
    ))


class ExcelExporter:
    """
    Generates multi-sheet Excel workbook with prediction results.
    Uses openpyxl with dark green headers, alternating rows, conditional formatting.
    """

    def export(self, prediction_data: Dict, match_info: Dict) -> bytes:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        self._sheet_overview(wb, match_info, prediction_data)
        
        # New modular sections used to build the original tabs
        ws_dc = wb.create_sheet("1X2 & Double Chance")
        self._fill_1x2_dc_section(ws_dc, prediction_data, 1)
        
        ws_goals = wb.create_sheet("Goals Predictions")
        self._fill_goals_section(ws_goals, prediction_data, 1)
        
        ws_scores = wb.create_sheet("Score Predictions")
        self._fill_correct_scores_section(ws_scores, prediction_data, 1)
        
        ws_specials = wb.create_sheet("Special Markets")
        self._fill_specials_section(ws_specials, prediction_data, 1)
        
        ws_players = wb.create_sheet("Player Scorers")
        self._fill_player_scorers_section(ws_players, prediction_data, 1)
        
        self._sheet_raw_features(wb, prediction_data)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def export_batch(self, predictions: List[Dict], filters: Dict = None) -> bytes:
        """
        Export a batch of prediction results as a single Excel workbook.
        Provides a singular, comprehensive sheet with all relevant prediction data.
        """
        wb = openpyxl.Workbook()
        ws_sum = wb.active
        ws_sum.title = "Batch Summary"

        # ── Header Row (Comprehensive Summary) ────────────────────────────
        headers = [
            "Date", "League", "Home", "Away",
            "Home Win %", "Draw %", "Away Win %",
            "DC 1X %", "DC 12 %", "DC X2 %",
            "BTTS Yes %", "BTTS No %",
            "Over 2.5 %", "Top Score", "Top Score %",
            "Even %", "Odd %",
            "Goal Both Halves %",
            "0-15 %", "16-30 %", "31-45 %", "46-60 %", "61-75 %", "76-90 %", "90+ %",
            "Score Draw %"
        ]
        _header_row(ws_sum, 1, headers, freeze=True)

        # ── Helper to extract prediction data safely ──────────────────────
        def _extract(pred_entry):
            pred = pred_entry.get("result", pred_entry)
            home = pred_entry.get("home_team", "Home")
            away = pred_entry.get("away_team", "Away")
            return pred, home, away

        # ── Data Rows ─────────────────────────────────────────────────────
        for i, pred_entry in enumerate(predictions, 2):
            if not pred_entry.get("success", True):
                _data_row(ws_sum, i, [
                    pred_entry.get("match_date", ""),
                    pred_entry.get("competition_name", ""),
                    pred_entry.get("home_team", "N/A"),
                    pred_entry.get("away_team", "N/A"),
                ] + ["ERROR"] * (len(headers) - 4), alt=i % 2 == 0)
                continue

            pred, home, away = _extract(pred_entry)
            mr = pred.get("match_result", {})
            dc = pred.get("double_chance", {})
            btts = pred.get("btts", {})
            ou = pred.get("total_goals", {}).get("over_under", {})
            cs = pred.get("correct_scores", [])
            top_score = cs[0].get("score", "—") if cs else "—"
            top_score_prob = cs[0].get("probability", 0) if cs else 0
            
            eo = pred.get("total_even_odd", {})
            gbh = pred.get("goal_both_halves", {})
            gi = pred.get("goal_interval", {})
            sd = pred.get("score_draw", {})

            row_data = [
                pred_entry.get("match_date", ""),
                pred_entry.get("competition_name", ""),
                home, away,
                mr.get("home", 0), mr.get("draw", 0), mr.get("away", 0),
                dc.get("1X", 0), dc.get("12", 0), dc.get("X2", 0),
                btts.get("yes", 0), btts.get("no", 0),
                ou.get("over_2_5", 0),
                top_score, top_score_prob,
                eo.get("even", 0), eo.get("odd", 0),
                gbh.get("yes", 0),
                gi.get("0-15", 0), gi.get("16-30", 0), gi.get("31-45", 0), 
                gi.get("46-60", 0), gi.get("61-75", 0), gi.get("76-90", 0), gi.get("90+", 0),
                sd.get("probability", 0)
            ]

            _pct_row(ws_sum, i, row_data, alt=i % 2 == 0)

        # ── Conditional Formatting & Cleanup ──────────────────────────────
        if len(predictions) > 0:
            end_r = len(predictions) + 1
            # Apply formatting to all probability columns (E through Z)
            # Home Win(E) to Top Score %(O) + Even%(P) to Score Draw%(Z)
            # Columns: E, F, G, H, I, J, K, L, M, O, P, Q, R, S, T, U, V, W, X, Y, Z
            prob_cols = [
                "E", "F", "G", "H", "I", "J", "K", "L", "M", "O",
                "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"
            ]
            for col_letter in prob_cols:
                _add_conditional_formatting(ws_sum, col_letter, 2, end_r)
        
        _autofit(ws_sum)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def _fill_match_sheet(self, ws, pred, home, away):
        """Helper to fill a single worksheet with detailed prediction data."""
        row = 1
        ws.merge_cells(f"A{row}:E{row}")
        title = ws.cell(row=row, column=1, value=f"⚽ {home} vs {away}")
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = HEADER_FILL
        title.alignment = Alignment(horizontal="center")
        row += 2

        # 1. 1X2 & Double Chance Section
        row = self._fill_1x2_dc_section(ws, pred, row)
        row += 2

        # 2. Goals Section (BTTS & Over/Under)
        row = self._fill_goals_section(ws, pred, row)
        row += 2

        # 3. Correct Scores Section
        row = self._fill_correct_scores_section(ws, pred, row)
        row += 2

        # 4. Special Markets (Even/Odd, Both Halves)
        row = self._fill_specials_section(ws, pred, row)
        row += 2

        # 5. Player Scorers
        row = self._fill_player_scorers_section(ws, pred, row)
        
        _autofit(ws)

    def _fill_1x2_dc_section(self, ws, pred, row):
        """Fills 1X2 and Double Chance section."""
        mr = pred.get("match_result", {})
        dc = pred.get("double_chance", {})

        _header_row(ws, row, ["Market", "Outcome", "Probability", "Decimal Odds", "Fractional Odds"])
        row += 1
        rows = [
            ("1X2", "Home Win (1)", mr.get("home", 0)),
            ("1X2", "Draw (X)", mr.get("draw", 0)),
            ("1X2", "Away Win (2)", mr.get("away", 0)),
            ("Double Chance", "1X (Home or Draw)", dc.get("1X", 0)),
            ("Double Chance", "12 (Home or Away)", dc.get("12", 0)),
            ("Double Chance", "X2 (Draw or Away)", dc.get("X2", 0)),
        ]
        start_row = row
        for market, outcome, prob in rows:
            decimal_odds = round(1 / prob, 2) if prob > 0 else "N/A"
            frac = self._to_fractional(prob)
            _pct_row(ws, row, [market, outcome, prob, decimal_odds, frac], alt=row % 2 == 0)
            row += 1
        
        _add_conditional_formatting(ws, "C", start_row, row - 1)
        return row

    def _fill_goals_section(self, ws, pred, row):
        """Fills BTTS and Over/Under Goals section."""
        btts = pred.get("btts", {})
        tg = pred.get("total_goals", {})
        ou = tg.get("over_under", {})

        _header_row(ws, row, ["Market", "Outcome", "Probability", "Decimal Odds"])
        row += 1
        rows_data = [
            ("BTTS", "Yes", btts.get("yes", 0)),
            ("BTTS", "No", btts.get("no", 0)),
            ("Total Goals", f"Total (predicted: {tg.get('predicted', 0):.2f})", None),
            ("Home Goals", f"Predicted: {tg.get('home_predicted', 0):.2f}", None),
            ("Away Goals", f"Predicted: {tg.get('away_predicted', 0):.2f}", None),
        ]
        # Over/Under rows
        for key, val in ou.items():
            label = key.replace("_", ".").replace("over.", "Over ").replace("under.", "Under ")
            rows_data.append(("Over/Under", label, val))

        start_row = row
        for i, (market, outcome, prob) in enumerate(rows_data):
            if prob is not None:
                odds = round(1 / prob, 2) if prob and prob > 0 else "N/A"
                _pct_row(ws, row, [market, outcome, prob, odds], alt=row % 2 == 0)
            else:
                _data_row(ws, row, [market, outcome, "—", "—"], alt=row % 2 == 0)
            row += 1

        _add_conditional_formatting(ws, "C", start_row, row - 1)
        return row

    def _fill_correct_scores_section(self, ws, pred, row):
        """Fills Correct Score predictions."""
        cs = pred.get("correct_scores", [])
        sd = pred.get("score_draw", {})

        _header_row(ws, row, ["Rank", "Score", "Probability", "Decimal Odds", "Category"])
        row += 1
        start_row = row
        for i, score in enumerate(cs[:10], 1): # Top 10 for batch
            prob = score.get("probability", 0)
            h = score.get("home_goals", 0)
            a = score.get("away_goals", 0)
            cat = "Home Win" if h > a else ("Draw" if h == a else "Away Win")
            odds = round(1 / prob, 1) if prob > 0 else "N/A"
            _pct_row(ws, row, [i, score.get("score"), prob, odds, cat], alt=row % 2 == 0)
            row += 1

        _add_conditional_formatting(ws, "C", start_row, row - 1)
        row += 1

        _header_row(ws, row, ["Score Draw", "Probability", "Likely Scores"])
        row += 1
        _data_row(ws, row, [
            "Score Draw",
            sd.get("probability", 0),
            ", ".join(sd.get("likely_scores", [])),
        ])
        ws.cell(row=row, column=2).number_format = "0.00%"
        row += 1
        return row

    def _fill_specials_section(self, ws, pred, row):
        """Fills even/odd, goal both halves, and intervals."""
        eo = pred.get("total_even_odd", {})
        gbh = pred.get("goal_both_halves", {})
        gi = pred.get("goal_interval", {})

        _header_row(ws, row, ["Market", "Outcome", "Probability", "Decimal Odds"])
        row += 1
        start_row = row
        specials = [
            ("Total Even/Odd", "Even", eo.get("even", 0)),
            ("Total Even/Odd", "Odd", eo.get("odd", 0)),
            ("Goal Both Halves", "Yes", gbh.get("yes", 0)),
            ("Goal Both Halves", "No", gbh.get("no", 0)),
        ]
        for i, (market, outcome, prob) in enumerate(specials):
            odds = round(1 / prob, 2) if prob > 0 else "N/A"
            _pct_row(ws, row, [market, outcome, prob, odds], alt=row % 2 == 0)
            row += 1

        row += 1
        # Goal intervals
        _header_row(ws, row, ["Goal Interval", "Probability", "Decimal Odds"])
        row += 1
        interval_start = row
        for i, (interval, prob) in enumerate(gi.items()):
            odds = round(1 / prob, 2) if prob > 0 else "N/A"
            _pct_row(ws, row, [interval, prob, odds], alt=row % 2 == 0)
            row += 1

        _add_conditional_formatting(ws, "C", start_row, start_row + 3)
        _add_conditional_formatting(ws, "B", interval_start, row - 1)
        return row

    def _fill_player_scorers_section(self, ws, pred, row):
        """Fills player scorers list."""
        players = pred.get("player_scorers", [])

        _header_row(ws, row, [
            "Rank", "Name", "Team", "#", "Position",
            "Goals (Season)", "xG/90", "Score Probability", "Decimal Odds"
        ])
        row += 1
        start_row = row
        for i, p in enumerate(players[:15], 1): # Top 15
            prob = p.get("probability", 0)
            odds = round(1 / prob, 2) if prob > 0 else "N/A"
            _pct_row(ws, row, [
                i,
                p.get("name"),
                p.get("team", "").capitalize(),
                p.get("jersey_no", ""),
                p.get("position", ""),
                p.get("goals_season", 0),
                p.get("xg_per90", 0),
                prob,
                odds,
            ], alt=row % 2 == 0)
            row += 1

        if len(players) > 0:
            _add_conditional_formatting(ws, "H", start_row, row - 1)
        return row

    # ─────────────────────────────────────────────────────────────────────────────
    # Sheet 1: Match Overview
    # ─────────────────────────────────────────────────────────────────────────────

    def _sheet_overview(self, wb, match_info: Dict, pred: Dict):
        ws = wb.create_sheet("Match Overview")
        row = 1

        ws.merge_cells(f"A{row}:F{row}")
        title = ws.cell(row=row, column=1, value="⚽ FOOTBALL PREDICTION REPORT")
        title.font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
        title.fill = PatternFill(start_color="0A0E1A", end_color="0A0E1A", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 30
        row += 1

        info = [
            ("Home Team", match_info.get("home_team", "N/A")),
            ("Away Team", match_info.get("away_team", "N/A")),
            ("Competition", match_info.get("competition", "N/A")),
            ("Match Date", match_info.get("date", "N/A")),
            ("Venue", match_info.get("venue", "N/A")),
            ("Model Confidence", f"{pred.get('model_confidence', 0):.1%}"),
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]
        for label, value in info:
            _data_row(ws, row, [label, value])
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

        row += 1
        # Home Lineup
        home_lineup = match_info.get("home_lineup", [])
        if home_lineup:
            _header_row(ws, row, ["Home XI — #", "Name", "Position", "Goals (Season)", "xG/90"])
            row += 1
            for i, p in enumerate(home_lineup[:11]):
                _data_row(ws, row, [
                    p.get("jersey_no", ""),
                    p.get("name", ""),
                    p.get("position", ""),
                    p.get("goals_season", 0),
                    f"{p.get('xg_per90', 0):.2f}",
                ], alt=i % 2 == 1)
                row += 1

        row += 1
        # Away Lineup
        away_lineup = match_info.get("away_lineup", [])
        if away_lineup:
            _header_row(ws, row, ["Away XI — #", "Name", "Position", "Goals (Season)", "xG/90"])
            row += 1
            for i, p in enumerate(away_lineup[:11]):
                _data_row(ws, row, [
                    p.get("jersey_no", ""),
                    p.get("name", ""),
                    p.get("position", ""),
                    p.get("goals_season", 0),
                    f"{p.get('xg_per90', 0):.2f}",
                ], alt=i % 2 == 1)
                row += 1

        _autofit(ws)


    # ─────────────────────────────────────────────────────────────────────────────
    # Sheet 7: Raw Features
    # ─────────────────────────────────────────────────────────────────────────────

    def _sheet_raw_features(self, wb, pred: Dict):
        ws = wb.create_sheet("Raw Feature Data")
        features = pred.get("features_used", {})

        _header_row(ws, 1, ["Feature Name", "Value", "Description"])
        descriptions = {
            "home_avg_goals_scored_5": "Home avg goals scored (last 5 matches)",
            "home_avg_goals_scored_10": "Home avg goals scored (last 10 matches)",
            "away_avg_goals_scored_5": "Away avg goals scored (last 5 matches)",
            "away_avg_goals_scored_10": "Away avg goals scored (last 10 matches)",
            "home_avg_goals_conceded_5": "Home avg goals conceded (last 5)",
            "away_avg_goals_conceded_5": "Away avg goals conceded (last 5)",
            "home_win_rate_10": "Home win rate (last 10 matches)",
            "away_win_rate_10": "Away win rate (last 10 matches)",
            "home_form_points": "Home form points (last 5)",
            "away_form_points": "Away form points (last 5)",
            "h2h_home_wins": "H2H home wins (last 5 meetings)",
            "h2h_draws": "H2H draws (last 5 meetings)",
            "h2h_away_wins": "H2H away wins (last 5 meetings)",
            "home_league_position": "Home team league position",
            "away_league_position": "Away team league position",
            "home_first_half_goals_avg": "Home avg 1st half goals (last 5)",
            "away_first_half_goals_avg": "Away avg 1st half goals (last 5)",
            "home_second_half_goals_avg": "Home avg 2nd half goals (last 5)",
            "away_second_half_goals_avg": "Away avg 2nd half goals (last 5)",
            "home_clean_sheet_rate": "Home clean sheet rate (last 5)",
            "away_clean_sheet_rate": "Away clean sheet rate (last 5)",
            "match_importance_weight": "Match importance multiplier",
            "home_avg_xg_per90": "Home avg xG per 90 (lineup)",
            "away_avg_xg_per90": "Away avg xG per 90 (lineup)",
        }

        for i, (feat, val) in enumerate(features.items(), 2):
            desc = descriptions.get(feat, "Engineered feature")
            _data_row(ws, i, [feat, round(float(val), 4) if isinstance(val, float) else val, desc], alt=i % 2 == 0)

        _autofit(ws)

    # ─────────────────────────────────────────────────────────────────────────────

    @staticmethod
    def _to_fractional(prob: float) -> str:
        if prob <= 0 or prob >= 1:
            return "N/A"
        decimal = 1 / prob
        numerator = decimal - 1
        # Simplify to reasonable fractional
        for denom in [1, 2, 4, 5, 8, 10, 20]:
            num = round(numerator * denom)
            if num > 0:
                return f"{num}/{denom}"
        return f"{round(numerator * 10)}/10"

    def _add_filters_sheet(self, wb, filters: Dict):
        """Creates a summary sheet for the filters applied to this export."""
        ws = wb.create_sheet("Applied Filters", index=0)
        
        # Header Row (Matching style)
        ws.merge_cells("A1:C1")
        cell = ws["A1"]
        cell.value = "APPLIED EXPORT FILTERS"
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 25

        # Also add border to the other merged cells in the header row
        for col in range(1, 4):
            ws.cell(row=1, column=col).border = THIN_BORDER

        # Data Rows
        labels = {
            "market": "Target Market",
            "min_prob": "Minimum Probability (%)",
            "date": "Export Date/Time",
            "count": "Matches Included"
        }
        
        # Convert values to display format
        data = {
            "market": str(filters.get("market", "any")).replace("_", " ").title(),
            "min_prob": f"{filters.get('min_prob', 0) * 100:.0f}%",
            "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "count": filters.get("count", 0)
        }

        row = 3
        for key, label in labels.items():
            lc = ws.cell(row=row, column=2, value=label)
            lc.font = Font(bold=True)
            lc.border = THIN_BORDER
            lc.fill = ALT_FILL if row % 2 == 0 else PatternFill(fill_type=None)
            
            vc = ws.cell(row=row, column=3, value=data.get(key))
            vc.border = THIN_BORDER
            vc.fill = ALT_FILL if row % 2 == 0 else PatternFill(fill_type=None)
            vc.alignment = Alignment(horizontal="right")
            
            row += 1
            
        _autofit(ws)
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 25


    def _add_batch_overview_sheet(self, wb, predictions: List[Dict]):
        """Creates a high-level summary sheet for the entire batch results."""
        ws = wb.create_sheet("Batch Overview", index=1 if "Applied Filters" in wb.sheetnames else 0)
        
        # Header Row (Matching style)
        ws.merge_cells("A1:C1")
        cell = ws["A1"]
        cell.value = "BATCH RESULTS OVERVIEW"
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 25

        # Also add border to the other merged cells in the header row
        for col in range(1, 4):
            ws.cell(row=1, column=col).border = THIN_BORDER

        # Calculate Metrics
        total = len(predictions)
        valid = [p for p in predictions if p.get("success", True)]
        failed = total - len(valid)
        
        home_wins = 0
        draws = 0
        away_wins = 0
        total_conf = 0
        
        for p_entry in valid:
            p = p_entry.get("result", p_entry)
            mr = p.get("match_result", {})
            total_conf += p_entry.get("model_confidence", 0)
            
            # Use favored logic
            max_v = max(mr.get("home", 0), mr.get("draw", 0), mr.get("away", 0))
            if max_v == mr.get("home"): home_wins += 1
            elif max_v == mr.get("draw"): draws += 1
            else: away_wins += 1

        avg_conf = (total_conf / len(valid)) if valid else 0
        
        # Data Rows
        labels = {
            "total": "Total Matches Processed",
            "successful": "Successful Predictions",
            "failed": "Failed/Error Matches",
            "home": "Home Wins Predicted",
            "draw": "Draws Predicted",
            "away": "Away Wins Predicted",
            "conf": "Average Model Confidence"
        }
        
        data = {
            "total": total,
            "successful": len(valid),
            "failed": failed,
            "home": f"{home_wins} ({home_wins/len(valid):.1%})" if valid else "0",
            "draw": f"{draws} ({draws/len(valid):.1%})" if valid else "0",
            "away": f"{away_wins} ({away_wins/len(valid):.1%})" if valid else "0",
            "conf": f"{avg_conf:.1%}"
        }

        row = 3
        for key, label in labels.items():
            lc = ws.cell(row=row, column=2, value=label)
            lc.font = Font(bold=True)
            lc.border = THIN_BORDER
            lc.fill = ALT_FILL if row % 2 == 0 else PatternFill(fill_type=None)
            
            vc = ws.cell(row=row, column=3, value=data.get(key))
            vc.border = THIN_BORDER
            vc.fill = ALT_FILL if row % 2 == 0 else PatternFill(fill_type=None)
            vc.alignment = Alignment(horizontal="right")
            
            if key == "failed" and failed > 0:
                vc.font = Font(color="FF0000", bold=True)
            row += 1
            
        _autofit(ws)
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 25
            

# Singleton
excel_exporter = ExcelExporter()
